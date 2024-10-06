import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
from typing import List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont
import io

from multivec.utils.base_format import ImageDocument, TextDocument


class CSVLoader:
    """
    Load CSV File, create visualizations, and highlight important information.
    """

    def __init__(self, file_path: str, output_dir: Optional[str] = None):
        self.file_path = Path(file_path)
        self.output_dir = Path(output_dir) if output_dir else Path(tempfile.mkdtemp())
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.df = self._load_csv()

    def _load_csv(self) -> pd.DataFrame:
        return pd.read_csv(self.file_path)

    def _create_excel_visualization(self) -> str:
        wb = Workbook()
        ws = wb.active

        # Write headers
        for col, header in enumerate(self.df.columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Write data
        for row, data in enumerate(self.df.values, start=2):
            for col, value in enumerate(data, start=1):
                ws.cell(row=row, column=col, value=value)

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        excel_path = self.output_dir / f"{self.file_path.stem}_visualization.xlsx"
        wb.save(excel_path)
        return str(excel_path)

    def _excel_to_image(self, excel_path: str) -> Image.Image:
        wb = Workbook()
        ws = wb.active
        img = XLImage(excel_path)
        ws.add_image(img, "A1")

        img_buffer = io.BytesIO()
        wb.save(img_buffer)
        img_buffer.seek(0)

        return Image.open(img_buffer)

    def _highlight_important_info(self, image: Image.Image) -> Image.Image:
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default()

        # Example: Highlight cells with values above the mean
        mean_value = self.df.select_dtypes(include=[np.number]).mean().mean()
        highlight_coords = []

        for row, data in enumerate(self.df.values, start=1):
            for col, value in enumerate(data, start=1):
                if isinstance(value, (int, float)) and value > mean_value:
                    x = col * 100  # Adjust based on your image size
                    y = row * 30  # Adjust based on your image size
                    highlight_coords.append((x, y))

        for x, y in highlight_coords:
            draw.rectangle([x, y, x + 90, y + 20], outline="red", width=2)

        return image

    def process(self) -> Tuple[List[TextDocument], List[ImageDocument]]:
        """
        Process the CSV file, create visualizations, and highlight important information.

        Returns:
            Tuple[List[TextDocument], List[ImageDocument]]: Lists of text and image documents.
        """
        text_docs = [
            TextDocument(
                content=self.df.to_csv(index=False),
                metadata={"type": "text", "format": "csv"},
                page_index=0,
            )
        ]

        excel_path = self._create_excel_visualization()
        excel_image = self._excel_to_image(excel_path)
        highlighted_image = self._highlight_important_info(excel_image)

        image_path = self.output_dir / f"{self.file_path.stem}_highlighted.png"
        highlighted_image.save(image_path)

        image_docs = [
            ImageDocument(
                content=str(image_path),
                metadata={
                    "type": "image",
                    "format": "png",
                    "description": "Highlighted CSV visualization",
                },
                page_index=0,
            )
        ]

        return text_docs, image_docs
